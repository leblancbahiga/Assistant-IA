"""Tests unitaires pour ToolRegistry + ToolExecutor + DocumentGenerator — Sprint 6 v2"""
import json
import tempfile
from pathlib import Path

from src.tools.registry import (
    ToolRegistry,
    ToolDefinition,
    ToolParameter,
    ToolExecutor,
    ToolResult,
)
from src.tools.document import (
    DocumentGenerator,
    DocumentSpec,
    DocFormat,
    DocSection,
)


# ──────────────────────────────────────────────────
# Fixtures helpers
# ──────────────────────────────────────────────────


def _make_tool(name="tool_a", category="document", desc="A tool"):
    return ToolDefinition(
        name=name,
        description=desc,
        category=category,
        parameters=[
            ToolParameter(name="input", type="str", description="Input param")
        ],
    )


def _make_doc(
    title="Rapport",
    sections=None,
    fmt=DocFormat.MARKDOWN,
    metadata=None,
):
    if sections is None:
        sections = [DocSection(title="Intro", content="Bonjour", level=1)]
    return DocumentSpec(
        title=title, format=fmt, sections=sections, metadata=metadata or {}
    )


# ──────────────────────────────────────────────────
# ToolRegistry tests
# ──────────────────────────────────────────────────


class TestToolRegistry:

    def test_register_and_get(self):
        reg = ToolRegistry()
        t = _make_tool("alpha")
        reg.register(t)
        assert reg.get("alpha") is t
        assert reg.get("nonexistent") is None

    def test_unregister(self):
        reg = ToolRegistry()
        reg.register(_make_tool("a"))
        assert reg.unregister("a") is True
        assert len(reg) == 0
        assert reg.unregister("ghost") is False

    def test_list_tools(self):
        reg = ToolRegistry()
        reg.register(_make_tool("a"))
        reg.register(_make_tool("b"))
        tools = reg.list_tools()
        assert len(tools) == 2
        names = {t.name for t in tools}
        assert names == {"a", "b"}

    def test_list_by_category(self):
        reg = ToolRegistry()
        reg.register(_make_tool("a", category="document"))
        reg.register(_make_tool("b", category="web"))
        reg.register(_make_tool("c", category="document"))
        docs = reg.list_by_category("document")
        assert len(docs) == 2
        assert {t.name for t in docs} == {"a", "c"}

    def test_search(self):
        reg = ToolRegistry()
        reg.register(_make_tool("code_runner", desc="Run code snippets"))
        reg.register(_make_tool("web_search", desc="Search the web"))
        reg.register(_make_tool("doc_gen", desc="Generate documents"))
        assert len(reg.search("code")) == 1
        assert len(reg.search("search")) == 1  # web_search (name matches)
        # "search" matches web_search (name) + doc_gen (no "search" in name/desc) = 1
        assert len(reg.search("xyz")) == 0

    def test_search_no_result(self):
        reg = ToolRegistry()
        reg.register(_make_tool("a"))
        assert reg.search("zzz") == []

    def test_to_llm_schema(self):
        reg = ToolRegistry()
        reg.register(_make_tool("t"))
        schemas = reg.to_llm_schema()
        assert len(schemas) == 1
        assert schemas[0]["name"] == "t"
        assert "parameters" in schemas[0]
        assert schemas[0]["parameters"]["type"] == "object"

    def test_load_from_file(self):
        """Charge depuis un fichier JSON temporaire."""
        data = {
            "tools": [
                {
                    "name": "t1",
                    "description": "d1",
                    "category": "c1",
                    "parameters": [],
                },
                {
                    "name": "t2",
                    "description": "d2",
                    "category": "c2",
                    "parameters": [
                        {"name": "p", "type": "str", "description": "desc"}
                    ],
                },
            ]
        }
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False
        ) as f:
            json.dump(data, f)
            path = f.name
        reg = ToolRegistry()
        count = reg.load_from_file(path)
        assert count == 2
        assert len(reg) == 2
        assert reg.get("t1") is not None
        assert reg.get("t2") is not None

    def test_save_and_load(self):
        reg = ToolRegistry()
        reg.register(_make_tool("alpha", category="code"))
        reg.register(_make_tool("beta", category="memory"))
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            path = f.name
        reg.save_to_file(path)

        reg2 = ToolRegistry()
        count = reg2.load_from_file(path)
        assert count == 2
        assert len(reg2) == 2
        assert reg2.get("alpha") is not None
        assert reg2.get("beta") is not None

    def test_len(self):
        reg = ToolRegistry()
        assert len(reg) == 0
        reg.register(_make_tool("x"))
        assert len(reg) == 1
        reg.register(_make_tool("y"))
        assert len(reg) == 2


# ──────────────────────────────────────────────────
# ToolExecutor tests
# ──────────────────────────────────────────────────


class TestToolExecutor:

    def test_execute_success(self):
        reg = ToolRegistry()
        reg.register(_make_tool("add"))
        executor = ToolExecutor(reg)
        executor.register_handler("add", lambda x, y: x + y)
        result = executor.execute("add", {"x": 2, "y": 3})
        assert result.success is True
        assert result.output == 5
        assert result.error is None
        assert result.duration_ms >= 0

    def test_execute_unknown_tool(self):
        reg = ToolRegistry()
        executor = ToolExecutor(reg)
        result = executor.execute("ghost", {})
        assert result.success is False
        assert "inconnu" in result.error.lower()

    def test_execute_no_handler(self):
        reg = ToolRegistry()
        reg.register(_make_tool("orphan"))
        executor = ToolExecutor(reg)
        result = executor.execute("orphan", {})
        assert result.success is False
        assert "handler" in result.error.lower()

    def test_execute_exception(self):
        reg = ToolRegistry()
        reg.register(_make_tool("explode"))
        executor = ToolExecutor(reg)
        executor.register_handler("explode", lambda: (_ for _ in ()).throw(ValueError("boom")))
        result = executor.execute("explode", {})
        assert result.success is False
        assert "boom" in result.error

    def test_execute_returns_duration(self):
        reg = ToolRegistry()
        reg.register(_make_tool("noop"))
        executor = ToolExecutor(reg)
        executor.register_handler("noop", lambda: None)
        result = executor.execute("noop", {})
        assert result.success is True
        assert isinstance(result.duration_ms, float)


# ──────────────────────────────────────────────────
# DocumentGenerator tests
# ──────────────────────────────────────────────────


class TestDocumentGenerator:

    def setup_method(self):
        self.gen = DocumentGenerator()

    # ── Validation ──

    def test_validate_valid(self):
        spec = _make_doc()
        assert self.gen.validate_spec(spec) == []

    def test_validate_no_title(self):
        spec = _make_doc(title="")
        errors = self.gen.validate_spec(spec)
        assert any("Titre requis" in e for e in errors)

    def test_validate_no_sections(self):
        spec = _make_doc(sections=[])
        errors = self.gen.validate_spec(spec)
        assert any("section" in e.lower() for e in errors)

    def test_validate_empty_section(self):
        spec = _make_doc(sections=[DocSection(title="", content="text")])
        errors = self.gen.validate_spec(spec)
        assert any("titre requis" in e.lower() for e in errors)

    def test_validate_bad_format(self):
        spec = _make_doc(fmt="csv")
        errors = self.gen.validate_spec(spec)
        assert any("format" in e.lower() for e in errors)

    # ── Génération Word (.docx) ──

    def test_generate_word(self):
        """Vérifie que le fichier Word est créé et non vide."""
        spec = _make_doc(
            title="Rapport Word",
            metadata={"Auteur": "NURU"},
            fmt=DocFormat.WORD,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self.gen.generate(spec, Path(tmpdir) / "test.docx")
            assert path.exists()
            assert path.stat().st_size > 0
            assert path.suffix == ".docx"

    def test_generate_word_content(self):
        """Vérifie que le contenu Word contient le titre."""
        spec = _make_doc(
            title="Test Contenu Word",
            sections=[DocSection("Chapitre 1", "Paragraphe test", level=1)],
            fmt=DocFormat.WORD,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self.gen.generate(spec, Path(tmpdir) / "out.docx")
            from docx import Document

            doc = Document(str(path))
            texts = [p.text for p in doc.paragraphs]
            assert any("Test Contenu Word" in t for t in texts)

    # ── Génération PDF ──

    def test_generate_pdf(self):
        """Vérifie que le fichier PDF est créé et non vide."""
        spec = _make_doc(
            title="Rapport PDF",
            metadata={"Date": "2025-01-01"},
            fmt=DocFormat.PDF,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self.gen.generate(spec, Path(tmpdir) / "test.pdf")
            assert path.exists()
            assert path.stat().st_size > 0
            assert path.suffix == ".pdf"

    def test_generate_pdf_starts_with_magic(self):
        """Le PDF commence par %PDF."""
        spec = _make_doc(fmt=DocFormat.PDF)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self.gen.generate(spec, Path(tmpdir) / "test.pdf")
            with open(path, "rb") as f:
                header = f.read(5)
            assert header == b"%PDF-"

    # ── Génération PowerPoint (.pptx) ──

    def test_generate_pptx(self):
        """Vérifie que le fichier PPTX est créé et non vide."""
        spec = _make_doc(
            title="Présentation PPTX",
            metadata={"Sujet": "NURU"},
            sections=[
                DocSection("Slide 1", "Contenu slide 1", level=1),
                DocSection("Slide 2", "Contenu slide 2", level=1),
            ],
            fmt=DocFormat.PPTX,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self.gen.generate(spec, Path(tmpdir) / "test.pptx")
            assert path.exists()
            assert path.stat().st_size > 0
            assert path.suffix == ".pptx"

    def test_generate_pptx_slide_count(self):
        """Le PPTX a 1 slide titre + N slides sections."""
        spec = _make_doc(
            title="Slides",
            sections=[
                DocSection("A", "a", 1),
                DocSection("B", "b", 1),
                DocSection("C", "c", 1),
            ],
            fmt=DocFormat.PPTX,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self.gen.generate(spec, Path(tmpdir) / "test.pptx")
            from pptx import Presentation

            prs = Presentation(str(path))
            assert len(prs.slides) == 4  # 1 titre + 3 sections

    # ── Génération Excel (.xlsx) ──

    def test_generate_xlsx(self):
        """Vérifie que le fichier Excel est créé et non vide."""
        spec = _make_doc(
            title="Tableau XLSX",
            metadata={"Version": "1.0"},
            sections=[
                DocSection("Ligne 1", "Valeur 1", 1),
                DocSection("Ligne 2", "Valeur 2", 2),
            ],
            fmt=DocFormat.XLSX,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self.gen.generate(spec, Path(tmpdir) / "test.xlsx")
            assert path.exists()
            assert path.stat().st_size > 0
            assert path.suffix == ".xlsx"

    def test_generate_xlsx_content(self):
        """Le titre apparaît dans la cellule A1."""
        spec = _make_doc(
            title="Titre Excel",
            sections=[DocSection("X", "Y", 1)],
            fmt=DocFormat.XLSX,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self.gen.generate(spec, Path(tmpdir) / "test.xlsx")
            from openpyxl import load_workbook

            wb = load_workbook(str(path))
            ws = wb.active
            assert ws["A1"].value == "Titre Excel"

    # ── Génération Markdown ──

    def test_generate_markdown(self):
        """Le fichier Markdown contient le titre en heading."""
        spec = _make_doc(
            title="Mon Rapport MD",
            metadata={"Auteur": "NURU"},
            fmt=DocFormat.MARKDOWN,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self.gen.generate(spec, Path(tmpdir) / "test.md")
            assert path.exists()
            content = path.read_text(encoding="utf-8")
            assert "# Mon Rapport MD" in content
            assert "**Auteur** : NURU" in content
            assert "Bonjour" in content

    # ── Utilitaires ──

    def test_generate_json(self):
        spec = _make_doc(title="JSON Test", fmt=DocFormat.WORD)
        result = self.gen.generate_json(spec)
        assert result["title"] == "JSON Test"
        assert result["format"] == "word"
        assert len(result["sections"]) == 1
        assert result["sections"][0]["title"] == "Intro"

    def test_generate_from_template(self):
        template = "Bonjour {{name}}, score: {{score}}."
        result = self.gen.generate_from_template(
            template, {"name": "NURU", "score": "95"}
        )
        assert result == "Bonjour NURU, score: 95."

    def test_template_missing_keys(self):
        template = "Bonjour {{name}}, projet {{missing}}."
        result = self.gen.generate_from_template(template, {"name": "NURU"})
        assert "NURU" in result
        assert "{{missing}}" in result  # pas remplacé

    def test_create_parent_dirs(self):
        """Le répertoire parent est créé automatiquement."""
        spec = _make_doc(fmt=DocFormat.MARKDOWN)
        with tempfile.TemporaryDirectory() as tmpdir:
            deep = Path(tmpdir) / "a" / "b" / "c" / "doc.md"
            path = self.gen.generate(spec, deep)
            assert path.exists()

    def test_generate_all_formats(self):
        """Tous les formats génèrent un fichier non vide."""
        with tempfile.TemporaryDirectory() as tmpdir:
            for fmt in DocFormat:
                ext_map = {
                    DocFormat.WORD: ".docx",
                    DocFormat.PDF: ".pdf",
                    DocFormat.PPTX: ".pptx",
                    DocFormat.XLSX: ".xlsx",
                    DocFormat.MARKDOWN: ".md",
                }
                spec = _make_doc(title=f"Test {fmt.value}", fmt=fmt)
                path = self.gen.generate(spec, Path(tmpdir) / f"test{ext_map[fmt]}")
                assert path.exists(), f"Format {fmt.value} n'a pas créé le fichier"
                assert path.stat().st_size > 0, f"Format {fmt.value} est vide"

    def test_unsupported_format_raises(self):
        """Un format inconnu lève ValueError."""
        spec = _make_doc(fmt="docx")  # string, pas DocFormat
        with tempfile.TemporaryDirectory() as tmpdir:
            try:
                # DocFormat("docx") would work, but we test with invalid
                from src.tools.document import DocumentSpec, DocFormat

                class FakeFormat(str):
                    pass

                # Force an invalid format value
                bad_spec = DocumentSpec(
                    title="X",
                    format=DocFormat.WORD,  # valid — we'll test the else branch differently
                    sections=[DocSection("a", "b")],
                )
                # Instead, test by monkeypatching
                bad_spec.format = "INVALID"  # type: ignore
                self.gen.generate(bad_spec, Path(tmpdir) / "out.md")
                assert False, "Should have raised ValueError"
            except (ValueError, AttributeError):
                pass  # expected
